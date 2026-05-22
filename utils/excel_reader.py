from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def get_data(file_path, sheet_name="Sheet1"):

        try:
            workbook = load_workbook(file_path)
            sheet = workbook[sheet_name]

            data = []

            # Read headers (first row)
            headers = [
                cell.value
                for cell in next(sheet.iter_rows(min_row=1, max_row=1))
            ]

            # Read remaining rows
            for row in sheet.iter_rows(min_row=2, values_only=True):

                # Skip empty rows
                if not any(row):
                    continue

                row_dict = dict(zip(headers, row))
                data.append(row_dict)

            return data

        except FileNotFoundError:
            raise Exception(f" Excel file not found at: {file_path}")

        except KeyError:
            raise Exception(f" Sheet '{sheet_name}' not found in Excel file")

        except Exception as e:
            raise Exception(f" Error reading Excel file: {e}")